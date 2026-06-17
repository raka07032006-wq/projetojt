import openpyxl

try:
    wb = openpyxl.load_workbook("Rekap 5R HRGA - MEI 2026.xlsx", read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    
    # Let's inspect the first sheet
    sheet = wb.active
    print("Active sheet name:", sheet.title)
    
    # Print first 10 rows
    for r in range(1, 15):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        print(f"Row {r}: {row_vals}")
        
except Exception as e:
    print("Error:", e)
